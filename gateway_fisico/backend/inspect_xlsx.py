from pathlib import Path
from openpyxl import load_workbook
import os

path = os.path.expanduser('~\\Desktop\\tsea_fallback_chart.xlsx')
if not Path(path).exists():
    print('file not found', path)
    raise SystemExit(1)

wb = load_workbook(path, data_only=True)
print('sheets:', wb.sheetnames)
ws = wb[wb.sheetnames[0]]
print('first sheet title:', ws.title)

for r in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=6, values_only=True):
    print(r)
